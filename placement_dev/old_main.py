from openpyxl.styles import Font, GradientFill
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Fill
import random
import datetime
import PIL
import time

def sleep_three():
    time.sleep(3)

from copy import deepcopy



# TODO
# Add F / M
# Special aid, hearing


d = datetime.datetime
start_position = 'A2'
end_position = 'M45'

class_name = 'tk21'
new = '_placeringar_new.xlsx'
book_name = class_name + new
print(book_name)



def sheet_styling(start_position, end_position):
    """ Adds sheet styling in a specific interval, made possible by 'cell coordinate'"""
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            cell_coordinate = cell.coordinate
            selected_cell = sheet[cell_coordinate]
            selected_cell.font = Font(size=10)


hall = ['Default']

classes = {
    'tk22': ['Adrian Kasum',
             'Airo Hussein Meruf',
             'Andrei Parlea',
             'August Spernjak',
             'Carl Lagerlöf',
             'Christoffer Timany',
             'Clara Åkerberg',
             'Damon Bassiri',
             'Dumitru-Andrei',
             'Emil Kivi',
             'Enes Mart',
             'John Nilsson',
             'Koppány Mátó',
             'Leia Åslin',
             'Linnéa Persson',
             'Ludvig Roth',
             'Luka Tanasic',
             'Metin Kamal',
             'Mirna Janat',
             'Mona Ali ',
             'Mustafa Al-Maswari',
             'Naomi Ståhlman',
             'Nicole Manda',
             'Nina Bové',
             'Sebastian Hållander',
             'Shaemaa Alhmidi',
             'Shaima Bodaka',
             'Vilgot Bergkvist'],
    'tk21': [
        'Anik Devnath',
        'Arina Ali',
        'Balraj Singh',
        'Dino Kunic',
        'Felix Magnusson',
        'Filippa Börjesson',
        'Hugo Fagerström',
        'John Widén Nordlund',
        'Jolie Ha',
        'Livia Ekhammar',
        'Ludvig Brandt',
        'Naomi Akumbu',
        'Nikolas Peterson',
        'Shucayb Said',
        'Wilhelm Olofsson'
    ]
}

book = load_workbook('tk21_placeringar.xlsx')
sheet = book.active

current_class = classes.get(class_name)
number_of_students = len(current_class)
rand_student = list(range(number_of_students))
random.shuffle(rand_student)
rr_test = rand_student


def add_all_images(start_position, end_position, image_path, file_exstention):
    """ Take in a start cell and a end cell, eg A2 to M12. The name of the
     student has to be exactly match the name of the image name. Copied from schoolsoft."""
    path = image_path
    file_ending = file_exstention
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            cell_coordinate = cell.coordinate
            selected_cell = sheet[cell_coordinate]
            for student in current_class:
                if student == cell.value:
                    img = Image('{}{}{}'.format(path, student, file_ending).strip())
                    print(path, student, file_ending)
                    img.height = img.height / 10
                    img.width = img.width / 10
                    sheet.add_image(img, cell_coordinate)


def create_sheets():
    for h in hall:
        book.create_sheet(h)


def remove_empty_sheets():
    pass


#    for sheet in book:
#        if sheet.title != 'Default':
#            if cell.value is None:
#                book.remove(sheet)

def gen_random_seats():
    print('Sheet: ', sheet.title)
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            if cell.value == 'Placering':
                for i in range(len(rand_student)):
                    cell.value = '\t' + current_class[int(rand_student[i])]
                    print('Placed student: ', current_class[int(rand_student[i])])
                    rand_student.pop(i)
                    break


def generate_seats_working(rand_student):
    book.active = 1
    r = deepcopy(rand_student)
    print('Sheet: ', sheet.title)
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            if cell.value == 'Placering':
                for i in range(len(r)):
                    cell.value = '\t' + current_class[int(r[i])]
                    print('Placed student: ', current_class[int(r[i])])
                    r.pop(i)
                    break



def swap_sheets(hall):
    """for h in hall:
        for i, s in enumerate(book.sheetnames):
            if s == h:
                break"""

#print(book.sheetnames)
book.active = 0
#print(book.active)
generate_seats_working(rand_student)




# generate_seats_working()
# add_all_images(start_position, end_position, 'student_images/tk22_images/', '.jpeg')

# Styling
# sheet_styling(start_position, end_position)

# Adding images
# add_all_images(start_position, end_position, 'te20d_img/', '.png')

# Create sheets
# create_sheets()

# Remove empty sheets
# remove_empty_sheets()

# create_sheets()


# Save
book.save('tk21_placeringar_new.xlsx')
book.close()
