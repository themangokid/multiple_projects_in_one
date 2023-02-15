from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Fill
import random
import datetime
import PIL

d = datetime.datetime

start_position = 'A2'
end_position = 'H7'
class_name = 'tk22'
new = '_placeringar_new.xlsx'


def sheet_styling(start_position, end_position):
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            print('Mod 2 confirmed.')
            cell_coordinate = cell.coordinate
            selected_cell = sheet[cell_coordinate]
            selected_cell.font = Font(size=23)


hall = ['100',
        '101',
        '102',
        '117',
        '208',
        '209',
        '212',
        '214',
        '216',
        '217',
        '218',
        '219',
        '304',
        '307',
        '315',
        '318',
        '319',
        '320',
        '321',
        '322',
        '408',
        '412',
        '417',
        '418',
        '420',
        '425']

classes = {
    'tk22': ['Adrian Kasum',
             'Airo Hussein Meruf',
             'Andrei Parlea',
             'August Spernjak',
             'Carl Lagerlöf',
             'Christoffer Timany',
             'Clara Åkerberg',
             'Damon Bassiri',
             'Dumitru-Andrei',
             'Emil Kivi',
             'Enes Mart',
             'John Nilsson',
             'Koppány Mátó',
             'Leia Åslin',
             'Linnéa Persson',
             'Ludvig Roth',
             'Luka Tanasic',
             'Metin Kamal',
             'Mirna Janat',
             'Mona Ali ',
             'Mustafa Al-Maswari',
             'Naomi Ståhlman',
             'Nicole Manda',
             'Nina Bové',
             'Sebastian Hållander',
             'Shaemaa Alhmidi',
             'Shaima Bodaka',
             'Vilgot Bergkvist']
}

book = load_workbook(class_name + new)
sheet = book.active
current_class = classes.get(class_name)

random_numbers_list = []
number_of_students = len(current_class)
rand_student = list(range(number_of_students))
random.shuffle(rand_student)

sheet_styling(start_position, end_position)

for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
    for length, cell in enumerate(row):
        if cell.value == 'Placering':
            for i in range(len(rand_student)):
                cell.value = current_class[int(rand_student[i])]
                rand_student.pop(i)
                break


def add_image():
    img = Image('abstract.png')
    sheet.add_image(img, 'A1')


def remove_empty_sheets():
    pass

def swap_sheet():
    for s in book:
        pass


def create_sheets():
    for h in hall:
        book.create_sheet(h)


book.save(class_name + new)
book.close()
