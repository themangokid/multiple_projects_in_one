from openpyxl.styles import Font, GradientFill
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Fill
import random
import datetime
import PIL
import pathlib

path = pathlib.Path

d = datetime.datetime

start_position = 'A2'
end_position = 'M12'


def sheet_styling(start_position, end_position):
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            cell_coordinate = cell.coordinate
            selected_cell = sheet[cell_coordinate]
            selected_cell.font = Font(size=10)


hall = ['100',
        '101',
        '102',
        '117',
        '208',
        '209',
        '212',
        '214',
        '216',
        '217',
        '218',
        '219',
        '304',
        '307',
        '315',
        '318',
        '319',
        '320',
        '321',
        '322',
        '408',
        '412',
        '417',
        '418',
        '420',
        '425']

classes = {
    'te20d': ['Adolfsson Vincent',
              'Alfredsson Sebastian',
              'Andersson Linus',
              'Bergholtz Nils',
              'Carpio Flores Mattias',
              'Hama Gareb Adam',
              'Hellström Petersson Theo',
              'Henriksson Viggo',
              'Hermansson Isak',
              'Hermansson Lukas',
              'Hultén Tristan',
              'Kaminski Artur',
              'Klein Alexander',
              'Korzepa Dominik',
              'Lindholm Matthias',
              'Lithner Eskil',
              'Lundgren Oscar',
              'Matic Viktor',
              'Miftaraj Albin',
              'Nilsson Tobias',
              'Olsson Filippa',
              'Pålsson Isak',
              'Rogö Albert',
              'Ruan Kenny',
              'Sajadi Arien',
              'Stamnes Oliver',
              'Stude Adrian',
              'Tholence Etienne',
              'Vadi Elliot',
              'Vazae Nina',
              'Vidos Leon',
              'Zahirovic Adam']
}

book = load_workbook('te20d_placeringar.xlsx')
sheet = book.active

print(book)
print(sheet)


def get_current_worksheet():
    for s in book:
        sheet = book[s]
        print('Now in sheet: ' + sheet.title)


current_class = classes.get('te20d')

random_numbers_list = []
number_of_students = len(current_class)
rand_student = list(range(number_of_students))
random.shuffle(rand_student)

# sheet_styling(start_position, end_position)

for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
    for length, cell in enumerate(row):
        if cell.value == 'Placering':
            for i in range(len(rand_student)):
                cell.value = current_class[int(rand_student[i])]
                rand_student.pop(i)
                break


def add_all_images(start_position, end_position):
    path = 'student_images/te20d_img/'
    file_ending = '.png'
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            cell_coordinate = cell.coordinate
            selected_cell = sheet[cell_coordinate]
            for student in current_class:
                if student == cell.value:
                    img = Image('{}{}{}'.format(path, student, file_ending).strip())
                    img.height = img.height / 10
                    img.width = img.width / 10
                    sheet.add_image(img, cell_coordinate)


def remove_empty_sheets():
    try:
        for h in hall:
            print(sheet)
            sheet.remove_sheet(h)
    except:
        print('Could not remove sheet')


def create_sheets():
    for h in hall:
        book.create_sheet(h)


get_current_worksheet()
#remove_empty_sheets()
# add_all_images(start_position, end_position)
# sheet_styling(start_position, end_position)


book.save('te20d_placeringar_new.xlsx')
book.close()
