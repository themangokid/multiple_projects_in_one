from openpyxl import Workbook, load_workbook
import random
import datetime
from copy import deepcopy
import os

book = load_workbook('placeringar.xlsx')

d = datetime.datetime
start_position = 'A2'
end_position = 'M45'


# TODO
# 1. Make a new loop function that counts the number of "Placering" and prints if the class fits the hall
# 2. See if there is a way to add photos,


def create_hall_sheets():
    """
    Generate halls for the school
    :return:
    """
    for h in hall:
        book.create_sheet(h)


def place_student(rand_student, sheet_title):
    sheet = sheet_title
    r = deepcopy(rand_student)
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            if cell.value == 'Placering':
                for i in range(len(r)):
                    cell.value = '\t' + current_class[int(r[i])]
                    print('Placed student: ', current_class[int(r[i])])
                    r.pop(i)
                    break


def select_sheet_and_populate(sheet_name):
    book.active = book.sheetnames.index(sheet_name)
    sheet_sent = book.active
    place_student(rand_student, sheet_sent)


def populate_seatings():
    sheet_names = book.sheetnames
    for s_name in sheet_names:
        select_sheet_and_populate(s_name)


def select_sheet_and_get_hall_size(sheet_name):
    book.active = book.sheetnames.index(sheet_name)
    sheet_sent = book.active
    get_halls_size(sheet_sent)


def get_hall_size_loop():
    sheet_names = book.sheetnames
    for s_name in sheet_names:
        select_sheet_and_get_hall_size(s_name)


def get_halls_size(sheet_title):
    sheet = sheet_title
    hall_size = 0
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            if cell.value == 'Placering':
                hall_size += 1
    print(sheet.title, 'Capacity:', hall_size)
    # if group_size > hall_size :
    #    print('Student class size too large ')
    return hall_size


def get_all_group_sizes():
    for id in class_ids:
        number_of_students = classes.get(id)
        print(id, len(number_of_students))
        return len(number_of_students)


def get_group_size(id):
    number_of_students = classes.get(id)
    print(id, len(number_of_students))
    return len(number_of_students)


def check_single_class_and_hall_capacity(student_class, hall_name):
    book.active = book.sheetnames.index(hall_name)
    sheet_sent = book.active

    student_class = get_group_size(student_class)
    hall_capacity = get_halls_size(sheet_sent)

    if student_class > hall_capacity:
        print('Students do not fit in {}'.format(hall_name))


hall = ['Default',
        'Ohio',
        'Utah',
        'Montana',
        'Texas',
        'Colorado',
        'Washington',
        'HLS Silent',
        'ALC Green Room',
        'ALC',
        'Californa',
        'Nevada',
        'Idaho',
        'Arizona',
        'Wyoming',
        'Oregon',
        'Alabama',
        'Florida',
        'India',
        'Alaska',
        'Nebraska',
        'Tennessee']

classes = {
    'ek20a': [
        'Agnes Pauli',
        'Albin Falkdahl',
        'Alex O" Brien',
        'Alexander Tviksta',
        'Alvin Mohlin',
        'Anton Lundgren',
        'Armin Rahimi',
        'Benjamin Jones',
        'Carl Just',
        'Christopher Holm',
        'Edwin W??hlander',
        'Elina Lang',
        'Elliot Ingemarsson',
        'Erik Norl??n',
        'Euan Hussey',
        'Felix Fogelholm',
        'Filippa Kinde',
        'Gustav Holmstedt',
        'Hugo Andersson',
        'Issac Ngo',
        'Jack Falk',
        'Liam Crespo Bergstr??m',
        'Lina Lundh',
        'Linn??a Zagerholm',
        'Madison Lind',
        'May Saetiao',
        'Niklas Dirlinger',
        'Noah Ginberg',
        'Simon Jansson',
        'Tim Eklund Bergfalk'
    ],

    'ek20b': [
        'Albin ??ystil?? Salo',
        'Alex Kj??ll',
        'Alexander Eriksen',
        'Alva Harbron',
        'Diana Shafazand',
        'Elliott Str??msten',
        'Fabian Hosseini',
        'Filippa Nilsson',
        'Greta Ringblom',
        'Ida Johansson',
        'Irnes Muric',
        'Jonathan Stephan Humble',
        'Julia Irissi',
        'Julius Hansson',
        'Klaudija Santic',
        'Leon Makolli',
        'Lina Utterhall',
        'Linnea Hilling',
        'Linn??a Setterberg',
        'Linus Feuk Bj??rkman',
        'Matilda Dufva',
        'Mohamed Mohammed',
        'Niclas Mattson',
        'Nilo Kontturi',
        'Nils Arnvik',
        'Nils Lyrmark',
        'Nina Tomicevic',
        'San Rahim',
        'Tolga ??nal',
        'Zonya Warasathaen'
    ],

    'ek21a': [
        'Albin Durvik',
        'Bazz Celik',
        'Brandon Ullman Carmona',
        'Clara Moerth',
        'Emilia Gustin',
        'Hasan Rayes',
        'Ibrahim Aden',
        'Jacob Jobe',
        'Jasmine Didari',
        'John Bjelke Hallander',
        'Kajsa Br??nnberg',
        'Melvin H??gg',
        'Melvin Riley',
        'Milton M??gi',
        'Milton Wallner',
        'Nathalie De Sousa',
        'Nima Sandstr??m Danesh',
        'Noel From',
        'Nour Al Kasem',
        'Oscar Dejevik',
        'Rasmus Halonen',
        'Rina Eman',
        'Sara Al-Jeboori',
        'Sara Holmgren',
        'Shilan Zarei Ghaleh',
        'Simon Zou',
        'Tuva Aronsson',
        'Zacharias Gustavsson'],

    'ek21b': [
        'Alexander Pamp',
        'Angelina H????st',
        'Anton Lundin',
        'Bj??rn Sk??ld',
        'Emelie Thyr??n',
        'Engla Sameholm',
        'Erik Tormodsson',
        'Hilda Hamon Dicksson',
        'Joakim Fritz',
        'Johan Lindvall',
        'Johanna Wiberg',
        'Kellie Emanuelsson',
        'Lara Omer',
        'Liam Grevsten',
        'Lorena Merida',
        'Lukas Barth',
        'Najma Abdi',
        'Nima Raeisi',
        'Oliver Persson',
        'Olle Bredow',
        'Philip Forsling',
        'Ramla Mohamed Ali',
        'Ruben Emericks',
        'Shawn Kahori',
        'Silvia Zandi',
        'Simon Alm',
        'Susanna Smith',
        'Tindra Bryfors'
    ],

    'ek22a': [
        'Agnes Strandberg',
        'Alexander Palo',
        'Angelica Svensson',
        'Anton Vesterlund',
        'Charlie Nordenhav',
        'Daniella Gnanzi',
        'Elicia Frejd',
        'Elin St??lnacke',
        'Emil Magnusson',
        'Emma Carlsson',
        'Fabian ??berg',
        'Helin Ahmed',
        'Inez Vallenius',
        'Johann??s Kallunki Nilsson',
        'Julia K??hlmark',
        'Kahlif Khulud',
        'Leo Dahl',
        'Lina Aslan',
        'Lucas Ander',
        'Maja Kihlbom',
        'Maximus Andersson S??rnes',
        'Mona Elkhateeb',
        'Noah Dahl',
        'Oliver Jedbratt',
        'Oskar Andersson',
        'Sofia Ekimova',
        'Sydney Paro',
        'Theo Olsson',
        'Wilmer Pettersson'
    ],

    'ek22b': [
        'Alexander Wehrle',
        'Amanda Olsson',
        'Anton Hansson',
        'Avin Salih',
        'Diana Glinkina',
        'Elin Hansson',
        'Elis Grahn',
        'Elliot Schweitz',
        'Emilia M??ller',
        'Fabian Sanner',
        'Filip Aslan',
        'Hanna Henriques',
        'HannaLouise ??gren',
        'Ida Hall Birkeland',
        'Jacob H??gberg',
        'Jakob R??djer',
        'Johanna Ivarsson',
        'John Sundqvist Lundstr??m',
        'Julia Stengarn',
        'Lara Adnan',
        'Leon Ekel??f',
        'Linus L??fnertz',
        'Maja Elsius',
        'Melvin Ramstr??m',
        'Nike Kyl??n',
        'Olle Stoldt',
        'Santiago Uranga Olvera',
        'Sara Mohamed',
        'Sixten Widborn',
        'Theo Borgqvist Touronen',
        'Tilda Alexandersson',
        'Tindra Rask'
    ],

    'na20': [
        'Adam Hassanie',
        'Assal Sadawi',
        'Daniel Khoda-Bakhsh',
        'Devin Kameli',
        'Elma Nylin',
        'Frida Svanberg',
        'Lisa Heath',
        'Melina Andreopoulos',
        'Nicole ??stlund',
        'Simon B????th',
        'Vilmer Karlsson',
        'Yusro Mohamed'
    ],

    'na21': [
        'Albert Vestergren Lindroth',
        'Alice Gink',
        'Aline Ndagijimana',
        'Benjamin Bozinov Hodzic',
        'Bianca Eboskog',
        'Bj??rn Stewart',
        'Ciel Jelvestam Petersson',
        'Filippa Wolving',
        'Jibran Hannani',
        'Julia Yver??s',
        'Lila Gebeyehu',
        'Lova Lundberg Ydrefalk',
        'Lucy Molback',
        'Madelene Bergstr??m',
        'Oleksandr Verbanov',
        'Robin Dolietis',
        'Robin Lindqvist',
        'Shuaib Salih',
        'Sofia Sj??kvist',
        'Sofia Toft??s',
        'Sophia Mayle',
        'Stina G??ransson',
        'Sumaya Tahlil',
        'Tabarek Al-Samkary',
        'TuvaLisa Carlsson',
        'Viola Zhang',
        'Wiggo Wads?? Alfredsson'
    ],

    'na22': [
        'Abdullah Dibo',
        'Ahmed Tarek',
        'Alice Dahlin',
        'Alva Andersson',
        'Amanda Bjurmalm',
        'Amelia Lundqvist',
        'Carl-Josef Melki',
        'Darin Omar',
        'Delia McNair',
        'Delnaz Yousef',
        'Emil Olsson',
        'Emma Anding',
        'Hanne Salih',
        'Imtiyaz Isse',
        'Isabelle Goy',
        'Jessica Sj??lin',
        'Jonathan Bylund',
        'Jonathan Sagdahl',
        'Lavia Hadi Karim',
        'Lilly Swaid',
        'Lorena Safradin',
        'Marijana Vlastic',
        'Molly Hillberg',
        'Nellie Larsson',
        'Rahma Abokor',
        'Rakhsha Manoranjitham Anagananthan',
        'Rami Al Robai',
        'Sipan Salih',
        'Stella Mod??n',
        'Zaineb Muhammed'
    ],

    'sa20a': [
        'Agnieszka Borkowska',
        'Alina Lindhe',
        'Amanda Strindehag',
        'Amina Kenjar',
        'Amna Sutrovic',
        'Axel Karlsson',
        'Chantal Hasselblad',
        'Ebba Honsberg',
        'Elin Andersson',
        'Elsa Andersson',
        'Embla Simonsson',
        'Emelie Larsson',
        'Emil Grundberg',
        'Emma Bjurmalm',
        'Hannes Lilja',
        'Jimi Riihinen',
        'Johanna Niklasson',
        'Klara Amn??s',
        'Kristian Bylyku',
        'Kristina Fougelberg Mortensen',
        'Lily-Louise Bowman',
        'Linn??a Pesonen',
        'Linn??a S??vetun',
        'Linn??a Selsvik',
        'Maja Engdahl',
        'M??ns Kennborn',
        'Minna Hall',
        'Nanok Ringqvist',
        'Nemo Lukman',
        'Sara Leoni Sandahl',
        'Tom Rustamzada'
    ],

    'sa20b': [
        'Albert Bjerner',
        'Alice Hillhammar',
        'Alva N??slund',
        'Anton Magnusson',
        'Aron Bodin',
        'Benjamin Vesovic Olsson',
        'Elias Tilk',
        'Elin G??ransson',
        'Emilia S??derstr??m',
        'Engla Ahlstrand',
        'Ethel Wir??n',
        'Jennifer Askn??s',
        'Klara Jaktling',
        'Liam Irandost',
        'M??ns Berggren',
        'Markus Solver',
        'Mathilda Ivarsson',
        'Maya Bevenhall',
        'Melissa Musumeci',
        'Nell Olsson',
        'Noah Lindroth',
        'Olivia Zyznowski',
        'Petter De Meere',
        'Selma Hahne',
        'Smilla Sonesson',
        'Tamara Spirkoska',
        'Thelma Bengtsson',
        'Theodor G??thberg',
        'William Poletar',
        'Wilmer Kvevik'
    ],

    'sa21a': [
        'Agnes Stahre',
        'Alva Cedergren',
        'Celine Sleiman',
        'Dina Pilica',
        'Elicia Salomonsson',
        'Elsa Hattina',
        'Emily DeMattia',
        'Frida Enwall',
        'Gustaf Grenabo',
        'Jeweriya Musa',
        'Johanna Sentayehu',
        'John Turitz',
        'Julia Szrajda',
        'Linda Ali',
        'Linus Lobrant',
        'Marcus Morisbak Olsson',
        'Matilda Reidal',
        'Maya Ryrl??n',
        'Mire ??str??m',
        'Moa Winberg',
        'Oscar N??tt',
        'Sallie Collins',
        'Saynab Mohamed',
        'Selma Wiberg',
        'Tiffany Djanfrouz Tell',
        'Wilner Warnhammar'
    ],

    'sa21b': [
        'Abdulqader Elmi',
        'Agnes R??nn??ng',
        'Ali Ahmad',
        'Elsa Beresford Hartwell',
        'Emna Demirovic Soltani',
        'Erik Lernfelt Natri',
        'Evelina Lev??n',
        'Ida Forsman',
        'Isabelle Bolteus',
        'Jocelyn Matonbila Troncoso',
        'Kasper Nilsson',
        'Leo Folkesson',
        'Linda Aziz',
        'Linn Hultin',
        'Maja ??hrn',
        'Nasir Muhammud',
        'Olivia Cromberger',
        'Suki Koehler',
        'Tova S??rensson',
        'Viktor Jansson'
    ],

    'sa22a': [
        'Albin Hillgard',
        'Alice Hago',
        'Anna-Belle Bengtsson',
        'Beverly Chesami',
        'Embla Hagstr??m Ringius',
        'Felicia Berntsson',
        'Hamdiah Ahmed Dankwah',
        'Helena Van Dongen',
        'Jenni ??stlund',
        'Johannes Frisfors',
        'Julia Davis',
        'Linnea Fagerlund',
        'Lisa Ekered',
        'Lisen Forsell',
        'Louise Kluvetasch',
        'Madeleine Gustavsson',
        'Max Clasborg',
        'Maya Sj??str??m Kumar',
        'Mina Al-Jabbari',
        'Nova Drakenbrandt',
        'Rubi Alvarado Guzman',
        'Sara Ben Jilali',
        'Sofia Siyad',
        'Svea Franz??n',
        'Tesnim Abdurhman',
        'Thea Kollberg',
        'Tilde Lennstr??m',
        'Vilma Ericsson'
    ],

    'sa22b': [
        'Adam Larsson',
        'Adilia Stenback',
        'Agnes Svensson',
        'Angelina Lucic Holmstr??m',
        'Axel Nilsson',
        'Edvin Rehn',
        'Emelie Sj??gren',
        'Felicia Andersson-Sarning',
        'Glorija Angjelovska',
        'Hanna Halldorf',
        'Ikram Hiloule',
        'Jacqueline Smith',
        'Linn Hohlf??lt',
        'Linn??a Ling??rdsson Jansson',
        'Lova Lund',
        'Max Mirkovic',
        'Melody Okpamen',
        'Moa Persson',
        'Nadja Holming',
        'Nicolina Karlsson',
        'Prescious Triumph',
        'Sabaa Musa',
        'Savannah Gudmundsson',
        'Tilda Lindqvist',
        'Vilda Westerblad',
        'Ville Andreasson',
        'Wilma Gustavsson',
        'Ylli Kelmendi',
        'Zahra Bahadji'
    ],

    'tk20': [
        'Adam Abunaj',
        'Adrian Bernsro',
        'Ahmed Salih',
        'Aisha Attar',
        'Albin Karlsteen',
        'Amitoz Berry',
        'Ariyen Gonimar',
        'Arvid Hermansson',
        'Arvid Kvocka',
        'Astrid Koczynski',
        'Edvin Rogonja',
        'Emil Essung',
        'Fatmeh Saif Eddin',
        'Golsa Masoudian',
        'Hugo Hagberg',
        'Josefin Johansson',
        'Kacper Strzelecki',
        'Linus Johansson',
        'Lucas Larking',
        'Markus Rynner',
        'Maryan Osman',
        'Nicolas Laine',
        'Pontus Bor??n',
        'Rajbir Singh',
        'Ramtin Abdollahzadeh',
        'Samuel Wideskott',
        'Sean Br??mster',
        'Simon Anding',
        'Tove Funksj??'
    ],

    'tk21': [
        'Anik Devnath',
        'Arina Ali',
        'Balraj Singh',
        'Dino Kunic',
        'Felix Magnusson',
        'Filippa B??rjesson',
        'Hugo Fagerstr??m',
        'John Wid??n Nordlund',
        'Jolie Ha',
        'Livia Ekhammar',
        'Ludvig Brandt',
        'Naomi Akumbu',
        'Nikolas Peterson',
        'Shucayb Said',
        'Wilhelm Olofsson'
    ],

    'tk22': [
        'Adrian Kasum',
        'Airo Hussein Meruf',
        'Andrei Parlea',
        'August Spernjak',
        'Carl Lagerl??f',
        'Christoffer Timany',
        'Clara ??kerberg',
        'Damon Bassiri',
        'Dumitru-Andrei Varzaru',
        'Emil Kivi Wikeby',
        'Enes Mart',
        'John Nilsson',
        'Leia ??slin',
        'Linn??a Persson',
        'Ludvig Roth',
        'Luka Tanasic',
        'Metin Kamal Mostafa',
        'Mirna Janat',
        'Mona Ali',
        'Mustafa Al-Maswari',
        'Naomi St??hlman',
        'Nicole Manda',
        'Nina Bov??',
        'Sebastian H??llander L????b',
        'Shaemaa Alhmidi',
        'Shaima Bodaka',
        'Test Person',
        'Vilgot Bergkvist',
        'Z??ti M??t??'
    ]
}

class_ids = ['ek20a',
             'ek20b',
             'ek21a',
             'ek21b',
             'ek22a',
             'ek22b',
             'na20',
             'na21',
             'na22',
             'sa20a',
             'sa20b',
             'sa21a',
             'sa21b',
             'sa22a',
             'sa22b',
             'tk20',
             'tk21',
             'tk22']

if __name__ == '__main__':
    # check_single_class_and_hall_capacity('ek20a', 'Ohio')
    # get_hall_size_loop()
    try:
        path = 'generated_placements'
        os.chdir('./generated_placements/')
    except:
        print('Could not create file')

    """:cvar
    
    NOTE NOTE NOTE NOTE 
    DOES NOT COMPUTER THE RIGHT CLASSES
    
    """
    run = input('Generate? ')
    os.chdir('./generated_placements/')
    if run == 'y':
        for c in class_ids:

            book = load_workbook('placeringar.xlsx')
            print(book.sheetnames)

            current_class = classes.get(c)
            number_of_students = len(current_class)
            rand_student = list(range(number_of_students))
            random.shuffle(rand_student)

            populate_seatings()
            today = str(datetime.date.today())

            book.save('{}_placeringar_{}.xlsx'.format(c, today))
            book.close()
