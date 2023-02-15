from openpyxl import Workbook, load_workbook
import random
import datetime
from copy import deepcopy
import os

book = load_workbook('placeringar.xlsx')

d = datetime.datetime
start_position = 'A2'
end_position = 'M45'


# TODO
# 1. Make a new loop function that counts the number of "Placering" and prints if the class fits the hall
# 2. See if there is a way to add photos,


def create_hall_sheets():
    """
    Generate halls for the school
    :return:
    """
    for h in hall:
        book.create_sheet(h)


def place_student(rand_student, sheet_title):
    sheet = sheet_title
    # r = deepcopy(rand_student)
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            if cell.value == 'Placering':
                for i in range(len(rand_student)):
                    cell.value = '\t' + current_class[int(rand_student[i])]
                    # print('Placed student: ', current_class[int(rand_student[i])])
                    rand_student.pop(i)
                    break


def select_sheet_and_populate(sheet_name):
    """
    Selects sheets and then place student with 'place student'
    :param sheet_name:
    :return:
    """
    book.active = book.sheetnames.index(sheet_name)
    sheet_sent = book.active
    # print(sheet_sent.title)
    place_student(rand_student, sheet_sent)


def populate_seatings():
    """
        Coupled with select_sheet_and_populate
    """
    sheet_names = book.sheetnames
    for s_name in sheet_names:
        select_sheet_and_populate(s_name)


def select_sheet_and_get_hall_size(sheet_name):
    """
    Selects sheets and then place student with 'place student'
    :param sheet_name:
    :return:
    """
    book.active = book.sheetnames.index(sheet_name)
    sheet_sent = book.active
    get_halls_size(sheet_sent)


def get_hall_size_loop():
    """
        Coupled with select_sheet_and_populate
    """
    sheet_names = book.sheetnames
    for s_name in sheet_names:
        select_sheet_and_get_hall_size(s_name)


"""
    Positions that swaps "Placering" for student name
"""


def get_halls_size(sheet_title):
    """
        Bug: Does not swap sheets
    :return:
    """
    sheet = sheet_title
    hall_size = 0
    for row in sheet['{}'.format(start_position):'{}'.format(end_position)]:
        for length, cell in enumerate(row):
            if cell.value == 'Placering':
                hall_size += 1
    print(sheet.title, 'Capacity:', hall_size)
    return hall_size


"""
    Data used to generate seating plans
    Halls
    Classess = students and classname
    Class_ids is used to get the list of example 'tk22'
"""

hall = ['Default',
        'Ohio',
        'Utah',
        'Montana',
        'Texas',
        'Colorado',
        'Washington',
        'HLS Silent',
        'ALC Green Room',
        'ALC',
        'Californa',
        'Nevada',
        'Idaho',
        'Arizona',
        'Wyoming',
        'Oregon',
        'Alabama',
        'Florida',
        'India',
        'Alaska',
        'Nebraska',
        'Tennessee']

classes = {
    'ek20a': [
        'Agnes Pauli',
        'Albin Falkdahl',
        'Alex O" Brien',
        'Alexander Tviksta',
        'Alvin Mohlin',
        'Anton Lundgren',
        'Armin Rahimi',
        'Benjamin Jones',
        'Carl Just',
        'Christopher Holm',
        'Edwin Wåhlander',
        'Elina Lang',
        'Elliot Ingemarsson',
        'Erik Norlén',
        'Euan Hussey',
        'Felix Fogelholm',
        'Filippa Kinde',
        'Gustav Holmstedt',
        'Hugo Andersson',
        'Issac Ngo',
        'Jack Falk',
        'Liam Crespo Bergström',
        'Lina Lundh',
        'Linnéa Zagerholm',
        'Madison Lind',
        'May Saetiao',
        'Niklas Dirlinger',
        'Noah Ginberg',
        'Simon Jansson',
        'Tim Eklund Bergfalk'
    ],

    'ek20b': [
        'Albin Öystilä Salo',
        'Alex Kjäll',
        'Alexander Eriksen',
        'Alva Harbron',
        'Diana Shafazand',
        'Elliott Strömsten',
        'Fabian Hosseini',
        'Filippa Nilsson',
        'Greta Ringblom',
        'Ida Johansson',
        'Irnes Muric',
        'Jonathan Stephan Humble',
        'Julia Irissi',
        'Julius Hansson',
        'Klaudija Santic',
        'Leon Makolli',
        'Lina Utterhall',
        'Linnea Hilling',
        'Linnéa Setterberg',
        'Linus Feuk Björkman',
        'Matilda Dufva',
        'Mohamed Mohammed',
        'Niclas Mattson',
        'Nilo Kontturi',
        'Nils Arnvik',
        'Nils Lyrmark',
        'Nina Tomicevic',
        'San Rahim',
        'Tolga Önal',
        'Zonya Warasathaen'
    ],

    'ek21a': [
        'Albin Durvik',
        'Bazz Celik',
        'Brandon Ullman Carmona',
        'Clara Moerth',
        'Emilia Gustin',
        'Hasan Rayes',
        'Ibrahim Aden',
        'Jacob Jobe',
        'Jasmine Didari',
        'John Bjelke Hallander',
        'Kajsa Brännberg',
        'Melvin Hägg',
        'Melvin Riley',
        'Milton Mägi',
        'Milton Wallner',
        'Nathalie De Sousa',
        'Nima Sandström Danesh',
        'Noel From',
        'Nour Al Kasem',
        'Oscar Dejevik',
        'Rasmus Halonen',
        'Rina Eman',
        'Sara Al-Jeboori',
        'Sara Holmgren',
        'Shilan Zarei Ghaleh',
        'Simon Zou',
        'Tuva Aronsson',
        'Zacharias Gustavsson'],

    'ek21b': [
        'Alexander Pamp',
        'Angelina Hööst',
        'Anton Lundin',
        'Björn Sköld',
        'Emelie Thyrén',
        'Engla Sameholm',
        'Erik Tormodsson',
        'Hilda Hamon Dicksson',
        'Joakim Fritz',
        'Johan Lindvall',
        'Johanna Wiberg',
        'Kellie Emanuelsson',
        'Lara Omer',
        'Liam Grevsten',
        'Lorena Merida',
        'Lukas Barth',
        'Najma Abdi',
        'Nima Raeisi',
        'Oliver Persson',
        'Olle Bredow',
        'Philip Forsling',
        'Ramla Mohamed Ali',
        'Ruben Emericks',
        'Shawn Kahori',
        'Silvia Zandi',
        'Simon Alm',
        'Susanna Smith',
        'Tindra Bryfors'
    ],

    'ek22a': [
        'Agnes Strandberg',
        'Alexander Palo',
        'Angelica Svensson',
        'Anton Vesterlund',
        'Charlie Nordenhav',
        'Daniella Gnanzi',
        'Elicia Frejd',
        'Elin Stålnacke',
        'Emil Magnusson',
        'Emma Carlsson',
        'Fabian Åberg',
        'Helin Ahmed',
        'Inez Vallenius',
        'Johannés Kallunki Nilsson',
        'Julia Köhlmark',
        'Kahlif Khulud',
        'Leo Dahl',
        'Lina Aslan',
        'Lucas Ander',
        'Maja Kihlbom',
        'Maximus Andersson Sörnes',
        'Mona Elkhateeb',
        'Noah Dahl',
        'Oliver Jedbratt',
        'Oskar Andersson',
        'Sofia Ekimova',
        'Sydney Paro',
        'Theo Olsson',
        'Wilmer Pettersson'
    ],

    'ek22b': [
        'Alexander Wehrle',
        'Amanda Olsson',
        'Anton Hansson',
        'Avin Salih',
        'Diana Glinkina',
        'Elin Hansson',
        'Elis Grahn',
        'Elliot Schweitz',
        'Emilia Möller',
        'Fabian Sanner',
        'Filip Aslan',
        'Hanna Henriques',
        'HannaLouise Ågren',
        'Ida Hall Birkeland',
        'Jacob Högberg',
        'Jakob Rödjer',
        'Johanna Ivarsson',
        'John Sundqvist Lundström',
        'Julia Stengarn',
        'Lara Adnan',
        'Leon Ekelöf',
        'Linus Löfnertz',
        'Maja Elsius',
        'Melvin Ramström',
        'Nike Kylén',
        'Olle Stoldt',
        'Santiago Uranga Olvera',
        'Sara Mohamed',
        'Sixten Widborn',
        'Theo Borgqvist Touronen',
        'Tilda Alexandersson',
        'Tindra Rask'
    ],

    'na20': [
        'Adam Hassanie',
        'Assal Sadawi',
        'Daniel Khoda-Bakhsh',
        'Devin Kameli',
        'Elma Nylin',
        'Frida Svanberg',
        'Lisa Heath',
        'Melina Andreopoulos',
        'Nicole Östlund',
        'Simon Bååth',
        'Vilmer Karlsson',
        'Yusro Mohamed'
    ],

    'na21': [
        'Albert Vestergren Lindroth',
        'Alice Gink',
        'Aline Ndagijimana',
        'Benjamin Bozinov Hodzic',
        'Bianca Eboskog',
        'Björn Stewart',
        'Ciel Jelvestam Petersson',
        'Filippa Wolving',
        'Jibran Hannani',
        'Julia Yverås',
        'Lila Gebeyehu',
        'Lova Lundberg Ydrefalk',
        'Lucy Molback',
        'Madelene Bergström',
        'Oleksandr Verbanov',
        'Robin Dolietis',
        'Robin Lindqvist',
        'Shuaib Salih',
        'Sofia Sjökvist',
        'Sofia Toftås',
        'Sophia Mayle',
        'Stina Göransson',
        'Sumaya Tahlil',
        'Tabarek Al-Samkary',
        'TuvaLisa Carlsson',
        'Viola Zhang',
        'Wiggo Wadsö Alfredsson'
    ],

    'na22': [
        'Abdullah Dibo',
        'Ahmed Tarek',
        'Alice Dahlin',
        'Alva Andersson',
        'Amanda Bjurmalm',
        'Amelia Lundqvist',
        'Carl-Josef Melki',
        'Darin Omar',
        'Delia McNair',
        'Delnaz Yousef',
        'Emil Olsson',
        'Emma Anding',
        'Hanne Salih',
        'Imtiyaz Isse',
        'Isabelle Goy',
        'Jessica Sjölin',
        'Jonathan Bylund',
        'Jonathan Sagdahl',
        'Lavia Hadi Karim',
        'Lilly Swaid',
        'Lorena Safradin',
        'Marijana Vlastic',
        'Molly Hillberg',
        'Nellie Larsson',
        'Rahma Abokor',
        'Rakhsha Manoranjitham Anagananthan',
        'Rami Al Robai',
        'Sipan Salih',
        'Stella Modén',
        'Zaineb Muhammed'
    ],

    'sa20a': [
        'Agnieszka Borkowska',
        'Alina Lindhe',
        'Amanda Strindehag',
        'Amina Kenjar',
        'Amna Sutrovic',
        'Axel Karlsson',
        'Chantal Hasselblad',
        'Ebba Honsberg',
        'Elin Andersson',
        'Elsa Andersson',
        'Embla Simonsson',
        'Emelie Larsson',
        'Emil Grundberg',
        'Emma Bjurmalm',
        'Hannes Lilja',
        'Jimi Riihinen',
        'Johanna Niklasson',
        'Klara Amnäs',
        'Kristian Bylyku',
        'Kristina Fougelberg Mortensen',
        'Lily-Louise Bowman',
        'Linnéa Pesonen',
        'Linnéa Sävetun',
        'Linnéa Selsvik',
        'Maja Engdahl',
        'Måns Kennborn',
        'Minna Hall',
        'Nanok Ringqvist',
        'Nemo Lukman',
        'Sara Leoni Sandahl',
        'Tom Rustamzada'
    ],

    'sa20b': [
        'Albert Bjerner',
        'Alice Hillhammar',
        'Alva Näslund',
        'Anton Magnusson',
        'Aron Bodin',
        'Benjamin Vesovic Olsson',
        'Elias Tilk',
        'Elin Göransson',
        'Emilia Söderström',
        'Engla Ahlstrand',
        'Ethel Wirén',
        'Jennifer Asknäs',
        'Klara Jaktling',
        'Liam Irandost',
        'Måns Berggren',
        'Markus Solver',
        'Mathilda Ivarsson',
        'Maya Bevenhall',
        'Melissa Musumeci',
        'Nell Olsson',
        'Noah Lindroth',
        'Olivia Zyznowski',
        'Petter De Meere',
        'Selma Hahne',
        'Smilla Sonesson',
        'Tamara Spirkoska',
        'Thelma Bengtsson',
        'Theodor Göthberg',
        'William Poletar',
        'Wilmer Kvevik'
    ],

    'sa21a': [
        'Agnes Stahre',
        'Alva Cedergren',
        'Celine Sleiman',
        'Dina Pilica',
        'Elicia Salomonsson',
        'Elsa Hattina',
        'Emily DeMattia',
        'Frida Enwall',
        'Gustaf Grenabo',
        'Jeweriya Musa',
        'Johanna Sentayehu',
        'John Turitz',
        'Julia Szrajda',
        'Linda Ali',
        'Linus Lobrant',
        'Marcus Morisbak Olsson',
        'Matilda Reidal',
        'Maya Ryrlén',
        'Mire Åström',
        'Moa Winberg',
        'Oscar Nätt',
        'Sallie Collins',
        'Saynab Mohamed',
        'Selma Wiberg',
        'Tiffany Djanfrouz Tell',
        'Wilner Warnhammar'
    ],

    'sa21b': [
        'Abdulqader Elmi',
        'Agnes Rönnäng',
        'Ali Ahmad',
        'Elsa Beresford Hartwell',
        'Emna Demirovic Soltani',
        'Erik Lernfelt Natri',
        'Evelina Levén',
        'Ida Forsman',
        'Isabelle Bolteus',
        'Jocelyn Matonbila Troncoso',
        'Kasper Nilsson',
        'Leo Folkesson',
        'Linda Aziz',
        'Linn Hultin',
        'Maja Öhrn',
        'Nasir Muhammud',
        'Olivia Cromberger',
        'Suki Koehler',
        'Tova Sörensson',
        'Viktor Jansson'
    ],

    'sa22a': [
        'Albin Hillgard',
        'Alice Hago',
        'Anna-Belle Bengtsson',
        'Beverly Chesami',
        'Embla Hagström Ringius',
        'Felicia Berntsson',
        'Hamdiah Ahmed Dankwah',
        'Helena Van Dongen',
        'Jenni Östlund',
        'Johannes Frisfors',
        'Julia Davis',
        'Linnea Fagerlund',
        'Lisa Ekered',
        'Lisen Forsell',
        'Louise Kluvetasch',
        'Madeleine Gustavsson',
        'Max Clasborg',
        'Maya Sjöström Kumar',
        'Mina Al-Jabbari',
        'Nova Drakenbrandt',
        'Rubi Alvarado Guzman',
        'Sara Ben Jilali',
        'Sofia Siyad',
        'Svea Franzén',
        'Tesnim Abdurhman',
        'Thea Kollberg',
        'Tilde Lennström',
        'Vilma Ericsson'
    ],

    'sa22b': [
        'Adam Larsson',
        'Adilia Stenback',
        'Agnes Svensson',
        'Angelina Lucic Holmström',
        'Axel Nilsson',
        'Edvin Rehn',
        'Emelie Sjögren',
        'Felicia Andersson-Sarning',
        'Glorija Angjelovska',
        'Hanna Halldorf',
        'Ikram Hiloule',
        'Jacqueline Smith',
        'Linn Hohlfält',
        'Linnéa Lingårdsson Jansson',
        'Lova Lund',
        'Max Mirkovic',
        'Melody Okpamen',
        'Moa Persson',
        'Nadja Holming',
        'Nicolina Karlsson',
        'Prescious Triumph',
        'Sabaa Musa',
        'Savannah Gudmundsson',
        'Tilda Lindqvist',
        'Vilda Westerblad',
        'Ville Andreasson',
        'Wilma Gustavsson',
        'Ylli Kelmendi',
        'Zahra Bahadji'
    ],

    'tk20': [
        'Adam Abunaj',
        'Adrian Bernsro',
        'Ahmed Salih',
        'Aisha Attar',
        'Albin Karlsteen',
        'Amitoz Berry',
        'Ariyen Gonimar',
        'Arvid Hermansson',
        'Arvid Kvocka',
        'Astrid Koczynski',
        'Edvin Rogonja',
        'Emil Essung',
        'Fatmeh Saif Eddin',
        'Golsa Masoudian',
        'Hugo Hagberg',
        'Josefin Johansson',
        'Kacper Strzelecki',
        'Linus Johansson',
        'Lucas Larking',
        'Markus Rynner',
        'Maryan Osman',
        'Nicolas Laine',
        'Pontus Borén',
        'Rajbir Singh',
        'Ramtin Abdollahzadeh',
        'Samuel Wideskott',
        'Sean Brömster',
        'Simon Anding',
        'Tove Funksjö'
    ],

    'tk21': [
        'Anik Devnath',
        'Arina Ali',
        'Balraj Singh',
        'Dino Kunic',
        'Felix Magnusson',
        'Filippa Börjesson',
        'Hugo Fagerström',
        'John Widén Nordlund',
        'Jolie Ha',
        'Livia Ekhammar',
        'Ludvig Brandt',
        'Naomi Akumbu',
        'Nikolas Peterson',
        'Shucayb Said',
        'Wilhelm Olofsson'
    ],

    'tk22': [
        'Adrian Kasum',
        'Airo Hussein Meruf',
        'Andrei Parlea',
        'August Spernjak',
        'Carl Lagerlöf',
        'Christoffer Timany',
        'Clara Åkerberg',
        'Damon Bassiri',
        'Dumitru-Andrei Varzaru',
        'Emil Kivi Wikeby',
        'Enes Mart',
        'John Nilsson',
        'Leia Åslin',
        'Linnéa Persson',
        'Ludvig Roth',
        'Luka Tanasic',
        'Metin Kamal Mostafa',
        'Mirna Janat',
        'Mona Ali',
        'Mustafa Al-Maswari',
        'Naomi Ståhlman',
        'Nicole Manda',
        'Nina Bové',
        'Sebastian Hållander Lööb',
        'Shaemaa Alhmidi',
        'Shaima Bodaka',
        'Test Person',
        'Vilgot Bergkvist',
        'Zéti Mátó'
    ]
}

output_from_running = ['ek20b', 'ek21a', 'ek21b', 'ek22a', 'ek22b', 'na20', 'na21', 'na22', 'sa20a', 'sa20b', 'sa21a',
                       'sa21b', 'sa22a', 'sa22b', 'tk20', 'tk21', 'tk22']
class_ids = ['ek20a',
             'ek20b',
             'ek21a',
             'ek21b',
             'ek22a',
             'ek22b',
             'na20',
             'na21',
             'na22',
             'sa20a',
             'sa20b',
             'sa21a',
             'sa21b',
             'sa22a',
             'sa22b',
             'tk20',
             'tk21',
             'tk22']

if __name__ == '__main__':
    """
        1. If students does not fit in classroom, alert
        2. Generate classrooms either way, with warning
        3. Add files to placements folder
    """
    # get_hall_size_loop()
    try:
        path = 'generated_placements'
        os.chdir('./generated_placements/')
    except:
        print('Could not create file')

    # run = input('Generate? ')
    # if run == 'y':
    for c in class_ids:
        # book = load_workbook('../placeringar.xlsx')

        """:cvar

        NOTE NOTE NOTE NOTE 
        GET GROUP SIZE WORKS 

        """

        print(class_ids)
        current_class = classes.get(c)
        # print(current_class)
        # print('Class size:', len(current_class))
        number_of_students = len(current_class)
        rand_student = list(range(number_of_students))
        random.shuffle(rand_student)
        # print(rand_student)

        populate_seatings()

        today = str(datetime.date.today())

        book.save('{}_placeringar_{}.xlsx'.format(c, today))
        class_ids.remove(c)
        print(class_ids)
        book.close()