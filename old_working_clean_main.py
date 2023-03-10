from openpyxl import Workbook, load_workbook
import random
import datetime
from copy import deepcopy
import datetime
from concurrent.futures import ProcessPoolExecutor
from openpyxl import load_workbook
from time import perf_counter

# https://foss.heptapod.net/openpyxl/openpyxl/-/snippets/69 - Parallelization
# from functools import cache, lru_cache

begin_time = datetime.datetime.now()
print(datetime.datetime.now())
book = load_workbook('./generated_placements/Placeringar.xlsx', read_only=True, keep_vba=True, keep_links=True,
                     data_only=True)


def create_hall_sheets():
    for h in hall:
        book.create_sheet(h)


def place_student(rand_student, sheet_title):
    sheet = sheet_title
    r = rand_student
    for row in sheet[f'{start_position}':f'{end_position}']:
        for cell in row:
            if cell.value == 'Placering':
                for i in range(len(r)):
                    cell.value = '\t' + current_class[int(r[i])]
                    print('Placed student: ', current_class[int(r[i])])
                    r.pop(i)
                    break


def select_sheet_and_populate(sheet_name):
    book.active = book.sheetnames.index(sheet_name)
    sheet_sent = book.active
    place_student(rand_student, sheet_sent)


def populate_seatings():
    """
        Coupled with select_sheet_and populate
    """
    sheet_names = book.sheetnames
    for s_name in sheet_names:
        select_sheet_and_populate(s_name)


def all_in_one(rand_student, sheet_title):
    sheet_names = book.sheetnames
    for s_name in sheet_names:
        select_sheet_and_populate(s_name)
        book.active = book.sheetnames.index(sheet_names)
        sheet_sent = book.active
        sheet = sheet_title
        r = rand_student
        for row in sheet[f'{start_position}':f'{end_position}']:
            for cell in row:
                if cell.value == 'Placering':
                    for i in range(len(r)):
                        cell.value = '\t' + current_class[int(r[i])]
                        r.pop(i)
                        break


d = datetime.datetime
start_position = 'A3'
end_position = 'O17'

hall = ['Default',
        'Ohio',
        'Utah',
        'Montana',
        'Texas',
        'Colorado',
        'Washington',
        'HLS Silent',
        'ALC Green Room',
        'ALC',
        'Californa',
        'Nevada',
        'Idaho',
        'Arizona',
        'Wyoming',
        'Oregon',
        'Alabama',
        'Florida',
        'India',
        'Alaska',
        'Nebraska',
        'Tennessee']

classes = {
    'ek20a': [
        'Agnes Pauli',
        'Albin Falkdahl',
        'Alex O" Brien',
        'Alexander Tviksta',
        'Alvin Mohlin',
        'Anton Lundgren',
        'Armin Rahimi',
        'Benjamin Jones',
        'Carl Just',
        'Christopher Holm',
        'Edwin W??hlander',
        'Elina Lang',
        'Elliot Ingemarsson',
        'Erik Norl??n',
        'Euan Hussey',
        'Felix Fogelholm',
        'Filippa Kinde',
        'Gustav Holmstedt',
        'Hugo Andersson',
        'Issac Ngo',
        'Jack Falk',
        'Liam Crespo Bergstr??m',
        'Lina Lundh',
        'Linn??a Zagerholm',
        'Madison Lind',
        'May Saetiao',
        'Niklas Dirlinger',
        'Noah Ginberg',
        'Simon Jansson',
        'Tim Eklund Bergfalk'
    ],

    'ek20b': [
        'Albin ??ystil?? Salo',
        'Alex Kj??ll',
        'Alexander Eriksen',
        'Alva Harbron',
        'Diana Shafazand',
        'Elliott Str??msten',
        'Fabian Hosseini',
        'Filippa Nilsson',
        'Greta Ringblom',
        'Ida Johansson',
        'Irnes Muric',
        'Jonathan Stephan Humble',
        'Julia Irissi',
        'Julius Hansson',
        'Klaudija Santic',
        'Leon Makolli',
        'Lina Utterhall',
        'Linnea Hilling',
        'Linn??a Setterberg',
        'Linus Feuk Bj??rkman',
        'Matilda Dufva',
        'Mohamed Mohammed',
        'Niclas Mattson',
        'Nilo Kontturi',
        'Nils Arnvik',
        'Nils Lyrmark',
        'Nina Tomicevic',
        'San Rahim',
        'Tolga ??nal',
        'Zonya Warasathaen'
    ],

    'ek21a': [
        'Albin Durvik',
        'Bazz Celik',
        'Brandon Ullman Carmona',
        'Clara Moerth',
        'Emilia Gustin',
        'Hasan Rayes',
        'Ibrahim Aden',
        'Jacob Jobe',
        'Jasmine Didari',
        'John Bjelke Hallander',
        'Kajsa Br??nnberg',
        'Melvin H??gg',
        'Melvin Riley',
        'Milton M??gi',
        'Milton Wallner',
        'Nathalie De Sousa',
        'Nima Sandstr??m Danesh',
        'Noel From',
        'Nour Al Kasem',
        'Oscar Dejevik',
        'Rasmus Halonen',
        'Rina Eman',
        'Sara Al-Jeboori',
        'Sara Holmgren',
        'Shilan Zarei Ghaleh',
        'Simon Zou',
        'Tuva Aronsson',
        'Zacharias Gustavsson'],

    'ek21b': [
        'Alexander Pamp',
        'Angelina H????st',
        'Anton Lundin',
        'Bj??rn Sk??ld',
        'Emelie Thyr??n',
        'Engla Sameholm',
        'Erik Tormodsson',
        'Hilda Hamon Dicksson',
        'Joakim Fritz',
        'Johan Lindvall',
        'Johanna Wiberg',
        'Kellie Emanuelsson',
        'Lara Omer',
        'Liam Grevsten',
        'Lorena Merida',
        'Lukas Barth',
        'Najma Abdi',
        'Nima Raeisi',
        'Oliver Persson',
        'Olle Bredow',
        'Philip Forsling',
        'Ramla Mohamed Ali',
        'Ruben Emericks',
        'Shawn Kahori',
        'Silvia Zandi',
        'Simon Alm',
        'Susanna Smith',
        'Tindra Bryfors'
    ],

    'ek22a': [
        'Agnes Strandberg',
        'Alexander Palo',
        'Angelica Svensson',
        'Anton Vesterlund',
        'Charlie Nordenhav',
        'Daniella Gnanzi',
        'Elicia Frejd',
        'Elin St??lnacke',
        'Emil Magnusson',
        'Emma Carlsson',
        'Fabian ??berg',
        'Helin Ahmed',
        'Inez Vallenius',
        'Johann??s Kallunki Nilsson',
        'Julia K??hlmark',
        'Kahlif Khulud',
        'Leo Dahl',
        'Lina Aslan',
        'Lucas Ander',
        'Maja Kihlbom',
        'Maximus Andersson S??rnes',
        'Mona Elkhateeb',
        'Noah Dahl',
        'Oliver Jedbratt',
        'Oskar Andersson',
        'Sofia Ekimova',
        'Sydney Paro',
        'Theo Olsson',
        'Wilmer Pettersson'
    ],

    'ek22b': [
        'Alexander Wehrle',
        'Amanda Olsson',
        'Anton Hansson',
        'Avin Salih',
        'Diana Glinkina',
        'Elin Hansson',
        'Elis Grahn',
        'Elliot Schweitz',
        'Emilia M??ller',
        'Fabian Sanner',
        'Filip Aslan',
        'Hanna Henriques',
        'HannaLouise ??gren',
        'Ida Hall Birkeland',
        'Jacob H??gberg',
        'Jakob R??djer',
        'Johanna Ivarsson',
        'John Sundqvist Lundstr??m',
        'Julia Stengarn',
        'Lara Adnan',
        'Leon Ekel??f',
        'Linus L??fnertz',
        'Maja Elsius',
        'Melvin Ramstr??m',
        'Nike Kyl??n',
        'Olle Stoldt',
        'Santiago Uranga Olvera',
        'Sara Mohamed',
        'Sixten Widborn',
        'Theo Borgqvist Touronen',
        'Tilda Alexandersson',
        'Tindra Rask'
    ],

    'na20': [
        'Adam Hassanie',
        'Assal Sadawi',
        'Daniel Khoda-Bakhsh',
        'Devin Kameli',
        'Elma Nylin',
        'Frida Svanberg',
        'Lisa Heath',
        'Melina Andreopoulos',
        'Nicole ??stlund',
        'Simon B????th',
        'Vilmer Karlsson',
        'Yusro Mohamed'
    ],

    'na21': [
        'Albert Vestergren Lindroth',
        'Alice Gink',
        'Aline Ndagijimana',
        'Benjamin Bozinov Hodzic',
        'Bianca Eboskog',
        'Bj??rn Stewart',
        'Ciel Jelvestam Petersson',
        'Filippa Wolving',
        'Jibran Hannani',
        'Julia Yver??s',
        'Lila Gebeyehu',
        'Lova Lundberg Ydrefalk',
        'Lucy Molback',
        'Madelene Bergstr??m',
        'Oleksandr Verbanov',
        'Robin Dolietis',
        'Robin Lindqvist',
        'Shuaib Salih',
        'Sofia Sj??kvist',
        'Sofia Toft??s',
        'Sophia Mayle',
        'Stina G??ransson',
        'Sumaya Tahlil',
        'Tabarek Al-Samkary',
        'TuvaLisa Carlsson',
        'Viola Zhang',
        'Wiggo Wads?? Alfredsson'
    ],

    'na22': [
        'Abdullah Dibo',
        'Ahmed Tarek',
        'Alice Dahlin',
        'Alva Andersson',
        'Amanda Bjurmalm',
        'Amelia Lundqvist',
        'Carl-Josef Melki',
        'Darin Omar',
        'Delia McNair',
        'Delnaz Yousef',
        'Emil Olsson',
        'Emma Anding',
        'Hanne Salih',
        'Imtiyaz Isse',
        'Isabelle Goy',
        'Jessica Sj??lin',
        'Jonathan Bylund',
        'Jonathan Sagdahl',
        'Lavia Hadi Karim',
        'Lilly Swaid',
        'Lorena Safradin',
        'Marijana Vlastic',
        'Molly Hillberg',
        'Nellie Larsson',
        'Rahma Abokor',
        'Rakhsha Manoranjitham Anagananthan',
        'Rami Al Robai',
        'Sipan Salih',
        'Stella Mod??n',
        'Zaineb Muhammed'
    ],

    'sa20a': [
        'Agnieszka Borkowska',
        'Alina Lindhe',
        'Amanda Strindehag',
        'Amina Kenjar',
        'Amna Sutrovic',
        'Axel Karlsson',
        'Chantal Hasselblad',
        'Ebba Honsberg',
        'Elin Andersson',
        'Elsa Andersson',
        'Embla Simonsson',
        'Emelie Larsson',
        'Emil Grundberg',
        'Emma Bjurmalm',
        'Hannes Lilja',
        'Jimi Riihinen',
        'Johanna Niklasson',
        'Klara Amn??s',
        'Kristian Bylyku',
        'Kristina Fougelberg Mortensen',
        'Lily-Louise Bowman',
        'Linn??a Pesonen',
        'Linn??a S??vetun',
        'Linn??a Selsvik',
        'Maja Engdahl',
        'M??ns Kennborn',
        'Minna Hall',
        'Nanok Ringqvist',
        'Nemo Lukman',
        'Sara Leoni Sandahl',
        'Tom Rustamzada'
    ],

    'sa20b': [
        'Albert Bjerner',
        'Alice Hillhammar',
        'Alva N??slund',
        'Anton Magnusson',
        'Aron Bodin',
        'Benjamin Vesovic Olsson',
        'Elias Tilk',
        'Elin G??ransson',
        'Emilia S??derstr??m',
        'Engla Ahlstrand',
        'Ethel Wir??n',
        'Jennifer Askn??s',
        'Klara Jaktling',
        'Liam Irandost',
        'M??ns Berggren',
        'Markus Solver',
        'Mathilda Ivarsson',
        'Maya Bevenhall',
        'Melissa Musumeci',
        'Nell Olsson',
        'Noah Lindroth',
        'Olivia Zyznowski',
        'Petter De Meere',
        'Selma Hahne',
        'Smilla Sonesson',
        'Tamara Spirkoska',
        'Thelma Bengtsson',
        'Theodor G??thberg',
        'William Poletar',
        'Wilmer Kvevik'
    ],

    'sa21a': [
        'Agnes Stahre',
        'Alva Cedergren',
        'Celine Sleiman',
        'Dina Pilica',
        'Elicia Salomonsson',
        'Elsa Hattina',
        'Emily DeMattia',
        'Frida Enwall',
        'Gustaf Grenabo',
        'Jeweriya Musa',
        'Johanna Sentayehu',
        'John Turitz',
        'Julia Szrajda',
        'Linda Ali',
        'Linus Lobrant',
        'Marcus Morisbak Olsson',
        'Matilda Reidal',
        'Maya Ryrl??n',
        'Mire ??str??m',
        'Moa Winberg',
        'Oscar N??tt',
        'Sallie Collins',
        'Saynab Mohamed',
        'Selma Wiberg',
        'Tiffany Djanfrouz Tell',
        'Wilner Warnhammar'
    ],

    'sa21b': [
        'Abdulqader Elmi',
        'Agnes R??nn??ng',
        'Ali Ahmad',
        'Elsa Beresford Hartwell',
        'Emna Demirovic Soltani',
        'Erik Lernfelt Natri',
        'Evelina Lev??n',
        'Ida Forsman',
        'Isabelle Bolteus',
        'Jocelyn Matonbila Troncoso',
        'Kasper Nilsson',
        'Leo Folkesson',
        'Linda Aziz',
        'Linn Hultin',
        'Maja ??hrn',
        'Nasir Muhammud',
        'Olivia Cromberger',
        'Suki Koehler',
        'Tova S??rensson',
        'Viktor Jansson'
    ],

    'sa22a': [
        'Albin Hillgard',
        'Alice Hago',
        'Anna-Belle Bengtsson',
        'Beverly Chesami',
        'Embla Hagstr??m Ringius',
        'Felicia Berntsson',
        'Hamdiah Ahmed Dankwah',
        'Helena Van Dongen',
        'Jenni ??stlund',
        'Johannes Frisfors',
        'Julia Davis',
        'Linnea Fagerlund',
        'Lisa Ekered',
        'Lisen Forsell',
        'Louise Kluvetasch',
        'Madeleine Gustavsson',
        'Max Clasborg',
        'Maya Sj??str??m Kumar',
        'Mina Al-Jabbari',
        'Nova Drakenbrandt',
        'Rubi Alvarado Guzman',
        'Sara Ben Jilali',
        'Sofia Siyad',
        'Svea Franz??n',
        'Tesnim Abdurhman',
        'Thea Kollberg',
        'Tilde Lennstr??m',
        'Vilma Ericsson'
    ],

    'sa22b': [
        'Adam Larsson',
        'Adilia Stenback',
        'Agnes Svensson',
        'Angelina Lucic Holmstr??m',
        'Axel Nilsson',
        'Edvin Rehn',
        'Emelie Sj??gren',
        'Felicia Andersson-Sarning',
        'Glorija Angjelovska',
        'Hanna Halldorf',
        'Ikram Hiloule',
        'Jacqueline Smith',
        'Linn Hohlf??lt',
        'Linn??a Ling??rdsson Jansson',
        'Lova Lund',
        'Max Mirkovic',
        'Melody Okpamen',
        'Moa Persson',
        'Nadja Holming',
        'Nicolina Karlsson',
        'Prescious Triumph',
        'Sabaa Musa',
        'Savannah Gudmundsson',
        'Tilda Lindqvist',
        'Vilda Westerblad',
        'Ville Andreasson',
        'Wilma Gustavsson',
        'Ylli Kelmendi',
        'Zahra Bahadji'
    ],

    'tk20': [
        'Adam Abunaj',
        'Adrian Bernsro',
        'Ahmed Salih',
        'Aisha Attar',
        'Albin Karlsteen',
        'Amitoz Berry',
        'Ariyen Gonimar',
        'Arvid Hermansson',
        'Arvid Kvocka',
        'Astrid Koczynski',
        'Edvin Rogonja',
        'Emil Essung',
        'Fatmeh Saif Eddin',
        'Golsa Masoudian',
        'Hugo Hagberg',
        'Josefin Johansson',
        'Kacper Strzelecki',
        'Linus Johansson',
        'Lucas Larking',
        'Markus Rynner',
        'Maryan Osman',
        'Nicolas Laine',
        'Pontus Bor??n',
        'Rajbir Singh',
        'Ramtin Abdollahzadeh',
        'Samuel Wideskott',
        'Sean Br??mster',
        'Simon Anding',
        'Tove Funksj??'
    ],

    'tk21': [
        'Anik Devnath',
        'Arina Ali',
        'Balraj Singh',
        'Dino Kunic',
        'Felix Magnusson',
        'Filippa B??rjesson',
        'Hugo Fagerstr??m',
        'John Wid??n Nordlund',
        'Jolie Ha',
        'Livia Ekhammar',
        'Ludvig Brandt',
        'Naomi Akumbu',
        'Nikolas Peterson',
        'Shucayb Said',
        'Wilhelm Olofsson'
    ],

    'tk22': [
        'Adrian Kasum',
        'Airo Hussein Meruf',
        'Andrei Parlea',
        'August Spernjak',
        'Carl Lagerl??f',
        'Christoffer Timany',
        'Clara ??kerberg',
        'Damon Bassiri',
        'Dumitru-Andrei Varzaru',
        'Emil Kivi Wikeby',
        'Enes Mart',
        'John Nilsson',
        'Leia ??slin',
        'Linn??a Persson',
        'Ludvig Roth',
        'Luka Tanasic',
        'Metin Kamal Mostafa',
        'Mirna Janat',
        'Mona Ali',
        'Mustafa Al-Maswari',
        'Naomi St??hlman',
        'Nicole Manda',
        'Nina Bov??',
        'Sebastian H??llander L????b',
        'Shaemaa Alhmidi',
        'Shaima Bodaka',
        'Test Person',
        'Vilgot Bergkvist',
        'Z??ti M??t??'
    ]
}

class_ids = ['ek20a',
             'ek20b',
             'ek21a',
             'ek21b',
             'ek22a',
             'ek22b',
             'na20',
             'na21',
             'na22',
             'sa20a',
             'sa20b',
             'sa21a',
             'sa21b',
             'sa22a',
             'sa22b',
             'tk20',
             'tk21',
             'tk22']


def select_sheet_and_get_hall_size(sheet_name):
    book.active = book.sheetnames.index(sheet_name)
    sheet_sent = book.active
    return get_halls_size(sheet_sent)


def get_hall_size_loop():
    sheet_names = book.sheetnames
    for s_name in sheet_names:
        select_sheet_and_get_hall_size(s_name)


def get_halls_size(sheet_title):
    sheet = sheet_title
    hall_size = 0
    for row in sheet[f'{start_position}':f'{end_position}']:
        for cell in row:
            if cell.value == 'Placering':
                hall_size += 1
    print(sheet.title, 'Capacity:', hall_size)
    # if group_size > hall_size :
    #    print('Student class size too large ')
    return hall_size


def get_all_group_sizes():
    for id in class_ids:
        number_of_students = classes.get(id)
        print(id, len(number_of_students))
        return len(number_of_students)


def get_group_size(id):
    number_of_students = classes.get(id)
    print(id, len(number_of_students))
    return len(number_of_students)


def print_groups_sizes():
    for id in class_ids:
        number_of_students = classes.get(id)
        print(id, len(number_of_students))


def check_single_class_and_hall_capacity(student_class, hall_name):
    book.active = book.sheetnames.index(hall_name)
    sheet_sent = book.active

    student_class = get_group_size(student_class)
    hall_capacity = get_halls_size(sheet_sent)

    if student_class > hall_capacity:
        print(f'Students do not fit in {hall_name}')


def write_capacity_stats(file_name, content, type):
    try:
        with open(file_name + type, 'r+') as f:
            for r in content:
                f.writelines(r + '\n')
    except:
        print('Failed to' + c.ly + ' WRITE feedback to file')


def write_empty_capacity_stats(file_name, type):
    """
        Writes an empty file (feedback-herman_n.md)
    """
    try:
        with open('{}'.format(file_name + type), 'x+') as f:
            f.write('')
    except:
        print('Failed to' + c.ly + ' write to EMPTY file')


def capacity_check():
    sheet_names = book.sheetnames
    for s_name in sheet_names:
        hall_size = select_sheet_and_get_hall_size(s_name)
        for id in class_ids:
            current_class_list = classes.get(id)
            class_size = len(current_class_list)
            halls_checked_content = []
            if class_size > hall_size:
                print(id, f'({class_size})', 'cannot be in', s_name, f'({hall_size})', 'missing',
                      class_size - hall_size, 'seats')
                halls_checked_content.append(
                    '{}'.format(class_size, 'cannot be in', s_name, hall_size, 'missing', class_size - hall_size,
                                'seats'))

            if class_size <= hall_size:
                print(id, f'({class_size})', 'fits in', s_name, f'({hall_size})', '', hall_size - class_size, 'seats')
                halls_checked_content.append(
                    '{}'.format(class_size, 'fits in', s_name, hall_size, '', hall_size - class_size,
                                'seats'))

            write_empty_capacity_stats('capacity_check.md')
            write_capacity_stats('capacity_check.md')


def save_stats():
    with open('../multiple_projects_in_one/other/pages_changed_remote/stats.md', 'a+') as f:
        diff = datetime.datetime.now() - begin_time
        date = datetime.datetime.now()
        f.write(str(date))
        f.write(str(diff))


if __name__ == '__main__':
    # google drive link: https://drive.google.com/drive/folders/13moemQ2iI0DVrj1A8bpVafuFyleAINJv
    # print_groups_sizes()
    question_capa = input('Do you want to generate capacity check stats? y/n')
    if question_capa == 'y':
        capacity_check()
    question_2 = input('Do you want to generate seating plans? y/n')
    if question_2 == 'y':
        for c in class_ids:
            book = load_workbook('./generated_placements/Placeringar.xlsx')
            current_class = classes.get(c)
            number_of_students = len(current_class)
            rand_student = list(range(number_of_students))
            random.shuffle(rand_student)

            populate_seatings()

            today = str(datetime.date.today())

            # book.save(f'{c}_placeringar_{today}.xlsx')
            book.save(f'./generated_placements/{c}_placeringar_{today}.xlsx')
            book.close()

        print(datetime.datetime.now() - begin_time)
